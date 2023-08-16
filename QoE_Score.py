import get_netparam
import get_videoparm
import cal_qoe
import threading
import time
import scapy
from scapy.all import sniff
from scapy.layers.inet import Ether, TCP
from scapy.layers.inet6 import IPv6
import queue
import os
import glob
import common
import openpyxl
from datetime import datetime
import signal


# 全局变量
capture_duration = 2  # 捕获持续时间（秒）
pcap_counter = 0  # pcap文件计数器
capture_flag = True  # 捕获标志
filter_ip = "2001:250:1001:1044::9d"
apn_lock = threading.Lock()
apn_ready = 0


pcap_queue = queue.Queue()

def capture_pkt():
    global pcap_counter
    while capture_flag:
        # TODO 这里是用IP区分当前视频服务的包和其他包，在真实场景中，该如何获取本机ip或视频服务器对应的ip需进一步改进
        print('开始捕获网络包')
        host = "2001:250:1001:1044::9d"
        filter_str = f"ip6 host {host}"
        packet = sniff(filter=filter_str, timeout=capture_duration)
        print('网络包捕获完成')
        filtered_packets = [pkt for pkt in packet if IPv6 in pkt and pkt[IPv6].dst == filter_ip]
        pcap_filename = f'capture_{pcap_counter}.pcap'


        save_thread = threading.Thread(target=scapy.utils.wrpcap, args=(pcap_filename, filtered_packets))
        save_thread.start()
        pcap_queue.put(pcap_filename)
        print("no set")
        get_netparam.write_event.set()
        print("yes set")

        pcap_counter += 1


def cleanup():
    while capture_flag:
        time.sleep(60)  # 每一分钟执行一次清理
        # 定期清理pcap文件，保留最近的N个文件
        max_files = 20
        delete_file_num = 5
        pcap_files = glob.glob("*.pcap")
        pcap_files.sort(key=os.path.getmtime, reverse=True)

        if len(pcap_files) > max_files:
            files_to_delete = pcap_files[-delete_file_num:]
            for file in files_to_delete:
                os.remove(file)
                print(f"Deleted file: {file}")

def qoe_save_file(file_name):

    print("----------save start------------")
    workbook = openpyxl.load_workbook(file_name)
    # 选择要操作的工作表（这里假设工作表名称为 "qoe"）
    sheet = workbook["QoE"]
    # 获取当前工作表的最后一行行号
    last_row = sheet.max_row
    # 在下一行继续写入数据
    sheet.cell(row=last_row + 1, column=1, value=common.golbal_time_save[last_row - 2])
    sheet.cell(row=last_row + 1, column=2, value=common.golbal_qoe_save[last_row - 2])
    # 保存文件
    workbook.save(file_name)
    print("----------save end------------")
    
    

def QoE_th_1(m3u8_path,file_name):
    global pcap_counter, apn_ready
    save_count = 0

    # ^C处理，避免文件保存过程中程序意外终止损坏文件
    def signal_handler(sig, frame):
        print("Received SIGINT (Ctrl+C). Exiting gracefully.")
        qoe_save_thread.join()
        exit(0)

    signal.signal(signal.SIGINT, signal_handler)


    while capture_flag:
        count = 0
        # 等2s，等第一个数据包抓完
        if count == 0:
            time.sleep(2.5)
            count = 1
        pcap_name = pcap_queue.get()
        QoE = cal_qoe.QoEScore(pcap_name,m3u8_path)

        # 注意qoe文件写入保存时不可中断，使用signal处理中断信号
        common.golbal_qoe_save.append(QoE)
        # save_count = save_count + 1
        current_time = datetime.now()
        formatted_time = current_time.strftime("%Y-%m-%d %H:%M:%S")
        common.golbal_time_save.append(formatted_time)
        qoe_save_thread = threading.Thread(target=qoe_save_file, args=(file_name,))
        qoe_save_thread.daemon = True
        qoe_save_thread.start()
        qoe_save_thread.join()
        # print(save_count)


        # if common.golbal_qoeParamter[3] < 0:
        #     common.golbal_qoeParamter[7] = last_qoe
            # QoE = last_qoe
        # last_qoe = common.golbal_qoeParamter[7]
        common.golbal_delay = common.golbal_qoeParamter[6]
        common.golbal_lossrate = common.golbal_qoeParamter[3]
        apn_lock.acquire()
        apn_ready = 1
        apn_lock.release()

        pcap_queue.task_done()

def QoE_th():
    global pcap_counter
    while capture_flag:
        pcap_name = pcap_queue.get()
        QoE = cal_qoe.QoEScore(pcap_name)
        print('QoE = ', QoE)

        pcap_queue.task_done()


def main(m3u8_path):
    global capture_flag

    print("yes")
    while True:
        print(os.path.exists(m3u8_path))
        if os.path.exists(m3u8_path):
            break
    
    # create a excel to save qoe
    workbook = openpyxl.Workbook()
    # 创建一个工作表
    sheet = workbook.create_sheet("QoE")
    # 在第一行写入数据
    sheet.cell(row=1, column=1, value="time")
    sheet.cell(row=1, column=2, value="QoE")
    # 保存文件
    file_name = "QoE.xlsx"
    workbook.save(file_name)
    
    capture_thread = threading.Thread(target=capture_pkt)
    capture_thread.daemon = True
    capture_thread.start()

    qoe_thread = threading.Thread(target=QoE_th_1, args=(m3u8_path,file_name))
    qoe_thread.daemon = True
    qoe_thread.start()

    # TODO 记得恢复
    # cleanup_thread = threading.Thread(target=cleanup)
    # cleanup_thread.daemon = True
    # cleanup_thread.start()

    # time.sleep(100)
    # capture_flag = False




    capture_thread.join()
    qoe_thread.join()
    pcap_queue.join()
    # cleanup_thread.join()


if __name__ == "__main__":
    main()
